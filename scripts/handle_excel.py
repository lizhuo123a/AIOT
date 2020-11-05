# ======================
# -*- coding: utf-8 -*-
# @author:LiZhuo
# @time  :2020/10/12 9:27
# @email :358840393@qq.com
# 今天的你要比昨天的你更优秀！
# ======================
from openpyxl import load_workbook
from scripts.constants import TESTCASE_PAHT
import sys
sys.path.append("..")
class HandleExcel:
    def __init__(self,filename,sheetname=None):
        self.filename = filename
        self.sheetname = sheetname

    def get_cases(self):
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        list = []
        head_data = tuple(ws.iter_rows(max_row=1,values_only=True))[0]
        all_cases = tuple(ws.iter_rows(min_row=2,values_only=True))
        for case in all_cases:
            list.append(dict(zip(head_data,case)))
        return list

    def write_result(self,row,real,result):
        other_wb = load_workbook(self.filename)
        if self.sheetname is None:
            other_ws = other_wb.active
        else:
            other_ws = other_wb[self.sheetname]
        other_ws.cell(row=row,column=7,value=real)
        other_ws.cell(row=row,column=8,value=result)
        other_wb.save(self.filename)
if __name__ == '__main__':
    a = HandleExcel(TESTCASE_PAHT,"product").get_cases()
    print(a)
